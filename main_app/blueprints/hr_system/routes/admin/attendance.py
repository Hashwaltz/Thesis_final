from datetime import  datetime, time
from flask import render_template, redirect, url_for, flash, request, session, current_app, after_this_request, send_file
from flask_login import login_required, current_user
from sqlalchemy.exc import IntegrityError

from sqlalchemy import and_
import os
from werkzeug.utils import secure_filename
import re
from openpyxl import load_workbook


from main_app.helpers.decorators import admin_required
from main_app.models.hr_models import Employee, Department,  Attendance, LateComputation, extract_late_from_attendance
from main_app.models.user import User
from main_app.extensions import db
from main_app.helpers.functions import parse_date, allowed_file, ALLOWED_EXTENSIONS, UPLOAD_FOLDER

from main_app.blueprints.hr_system.routes.admin import hr_admin_bp



@hr_admin_bp.route('/attendance')
@login_required
@admin_required
def view_attendance():
    page = request.args.get('page', 1, type=int)
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    employee_filter = request.args.get('employee', '').strip()
    department_filter = request.args.get('department', '').strip()
    status_filter = request.args.get('status', '').strip()  

    # Base query
    query = Attendance.query.join(Employee).join(Employee.department)

    # Date filters
    try:
        if start_date and not end_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(Attendance.date == start_date_obj)

        elif end_date and not start_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(Attendance.date == end_date_obj)

        elif start_date and end_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Swap if user entered in reverse
            if end_date_obj < start_date_obj:
                start_date_obj, end_date_obj = end_date_obj, start_date_obj

            query = query.filter(
                and_(
                    Attendance.date >= start_date_obj,
                    Attendance.date <= end_date_obj
                )
            )
    except ValueError:
        pass

    # Employee and Department filters
    if employee_filter:
        query = query.filter(Attendance.employee_id == int(employee_filter))
    if department_filter:
        query = query.filter(Employee.department_id == int(department_filter))

    # ✅ Status filter
    if status_filter:
        query = query.filter(Attendance.status == status_filter)

    # Pagination
    attendances = query.order_by(Attendance.date.desc()).paginate(page=page, per_page=20, error_out=False)

    # Lists for dropdowns
    employees = Employee.query.filter_by(archived=False).all()
    departments = Department.query.order_by(Department.name.asc()).all()

    return render_template(
        'hr/admin/attendance/view_attendance.html',
        attendances=attendances,
        employees=employees,
        departments=departments,
        start_date=start_date,
        end_date=end_date,
        employee_filter=employee_filter,
        department_filter=department_filter,
        status_filter=status_filter  
    )



@hr_admin_bp.route('/import-attendance', methods=['GET', 'POST'])
@login_required
@admin_required
def import_attendance():
    """Handle file upload and parsing"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(request.url)
        
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                # Save temporarily
                upload_folder = 'uploads'
                os.makedirs(upload_folder, exist_ok=True)
                filepath = os.path.join(upload_folder, file.filename)
                file.save(filepath)
                
                # Parse Excel
                attendance_data = parse_attendance_excel(filepath)
                
                # Store in session for preview
                session['pending_attendance'] = attendance_data
                
                # Clean up
                os.remove(filepath)
                
                flash(f'Parsed {len(attendance_data)} records. Please review.', 'info')
                return redirect(url_for('hr_admin_bp.preview_attendance'))
                
            except Exception as e:
                flash(f'Error: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Invalid format. Upload .xlsx file', 'error')
            return redirect(request.url)
    
    return render_template('hr/admin/attendance/import_attendance.html')


@hr_admin_bp.route('/preview-attendance')
@login_required
@admin_required
def preview_attendance():
    """Show preview of parsed attendance data"""
    pending_data = session.get('pending_attendance')
    if not pending_data:
        flash('No data to preview', 'warning')
        return redirect(url_for('hr_admin_bp.import_attendance'))
    
    # 🔑 Sort by user_id first, then by date_display string (YYYY-MM-DD sorts correctly)
    sorted_data = sorted(pending_data, key=lambda x: (x['user_id'], x['date_display']))
    
    return render_template('hr/admin/attendance/preview_attendance.html', 
                         attendance_data=sorted_data)



@hr_admin_bp.route('/confirm-attendance', methods=['POST'])
@login_required
@admin_required
def confirm_attendance():
    """Insert attendance records to database"""
    from datetime import time as dt_time, datetime as dt_datetime
    
    pending_data = session.get('pending_attendance')
    if not pending_data:
        flash('No pending data to import', 'error')
        return redirect(url_for('hr_admin_bp.view_attendance'))
    
    count = 0
    errors = []
    
    for record in pending_data:
        try:
            # Convert ISO strings back to time objects
            time_in = dt_time.fromisoformat(record['time_in']) if record.get('time_in') else None
            time_out = dt_time.fromisoformat(record['time_out']) if record.get('time_out') else None
            
            # 🔑 KEY FIX: If time_in exists but no time_out, set default to 5:00 PM
            if time_in and not time_out:
                time_out = dt_time(17, 0)  # 5:00 PM
            
            # 🔑 KEY FIX: Set status based on whether employee actually clocked in
            if time_in:
                status = "Present"
            else:
                status = "Absent"  # No time_in = Absent
            
            # Parse ISO date string back to Python date object
            date_value = record.get('date')
            if isinstance(date_value, str):
                date_obj = dt_datetime.strptime(date_value, '%Y-%m-%d').date()
            else:
                date_obj = date_value
            
            # Match Excel User ID to Employee.id (primary key integer)
            try:
                employee_id = int(record['user_id'])
                employee = Employee.query.get(employee_id)
            except (ValueError, TypeError):
                employee = None
            
            if not employee:
                errors.append(f"❌ Employee ID {record['user_id']} ({record.get('employee_name')}) not found")
                continue
            
            # Check if record already exists
            existing = Attendance.query.filter_by(
                employee_id=employee.id,
                date=date_obj
            ).first()
            
            if existing:
                # Update existing record
                existing.time_in = time_in
                existing.time_out = time_out
                existing.status = status  # ✅ Use computed status
                existing.remarks = "Updated via import" if not existing.remarks else existing.remarks
            else:
                # Create new record
                attendance = Attendance(
                    employee_id=employee.id,
                    date=date_obj,
                    time_in=time_in,
                    time_out=time_out,
                    status=status,  # ✅ "Present" or "Absent"
                    remarks="Imported from Excel"
                )
                db.session.add(attendance)
            
            count += 1
            
        except Exception as e:
            import traceback
            errors.append(f"❌ {record.get('employee_name')}: {str(e)}")
            if current_app.debug:
                print(f"DEBUG: {traceback.format_exc()}")
            db.session.rollback()
            continue
    
    try:
        db.session.commit()
        if errors:
            flash(f'⚠️ Partially imported: {count}/{len(pending_data)} records. {len(errors)} errors.', 'warning')
        else:
            flash(f'✅ Successfully imported {count} attendance records!', 'success')
    except Exception as e:
        db.session.rollback()
    finally:
        session.pop('pending_attendance', None)
    
    return redirect(url_for('hr_admin_bp.view_attendance'))


# Add to your routes file
@hr_admin_bp.app_template_filter('format_time_12')
def format_time_12(time_24):
    """Convert 24-hour time string (HH:MM) to 12-hour format (HH:MM AM/PM)"""
    try:
        from datetime import datetime
        dt = datetime.strptime(time_24, '%H:%M')
        return dt.strftime('%I:%M %p')
    except:
        return time_24


def parse_attendance_excel(filepath):
    """Parse attendance Excel file (supports both .xls and .xlsx)"""
    import re
    import calendar
    from datetime import datetime, date, time
    
    attendance_records = []
    current_employee = None
    processed_employees = set()
    
    # === FILE READING ===
    if filepath.endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        data = []
        for row_idx in range(ws.nrows):
            row = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row.append(dt)
                    except:
                        row.append(cell.value)
                else:
                    row.append(cell.value)
            data.append(row)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        data = [list(row) for row in ws.iter_rows(values_only=True)]
    
    # === EXTRACT DATE RANGE ===
    current_year = 2025
    current_month = 10
    
    for row in data:
        for cell in row:
            if cell and 'Attendance date' in str(cell):
                match = re.search(r'(\d{4})-(\d{2})', str(cell))
                if match:
                    current_year = int(match.group(1))
                    current_month = int(match.group(2))
                break
    
    days_in_month = calendar.monthrange(current_year, current_month)[1]
    
    # === FIND DATE COLUMNS (days 1 to days_in_month) ===
    date_columns = {}
    for row_idx, row in enumerate(data):
        for col_idx, cell in enumerate(row):
            if isinstance(cell, (int, float)) and 1 <= cell <= days_in_month:
                date_columns[col_idx] = int(cell)
    
    if not date_columns:
        raise Exception("No date columns found")
    
    # === PARSE EMPLOYEES ===
    for row_idx, row in enumerate(data):
        row_str = [str(cell) if cell is not None else '' for cell in row]
        row_text = ' '.join(row_str)
        
        # 🔑 Detect User ID row (employee header)
        if 'User ID:' in row_text:
            user_id = None
            name = ""
            department = ""
            
            for col_idx, cell in enumerate(row):
                if cell is not None:
                    cell_str = str(cell).strip()
                    if 'User ID:' in cell_str:
                        try:
                            user_id = int(re.search(r'User ID:\s*(\d+)', cell_str).group(1))
                        except:
                            if col_idx + 1 < len(row) and row[col_idx + 1]:
                                try:
                                    user_id = int(row[col_idx + 1])
                                except:
                                    pass
                    elif 'Name:' in cell_str:
                        name = cell_str.split('Name:')[-1].strip()
                    elif 'Department:' in cell_str:
                        department = cell_str.split('Department:')[-1].strip()
            
            # Skip if already processed this employee
            if user_id in processed_employees:
                continue
            processed_employees.add(user_id)
            
            # 🔑 LOOKUP EMPLOYEE BY PRIMARY KEY ID (Employee.id == Excel User ID)
            employee = None
            if user_id:
                try:
                    employee = Employee.query.get(int(user_id))
                except:
                    pass
            
            current_employee = {
                'user_id': user_id,
                'employee_name': employee.get_full_name() if employee else (name or f"User {user_id}"),
                'department': employee.department.name if employee and employee.department else department,
            }
            
            # 🔑 KEY FIX: Find the NEXT row that contains actual attendance times
            # Skip header rows, find the row with time data for this employee
            attendance_row = None
            for next_idx in range(row_idx + 1, min(row_idx + 10, len(data))):
                next_row = data[next_idx]
                # Check if this row has time data in date columns (not just day numbers)
                has_time_data = False
                for c in date_columns:
                    if c < len(next_row) and next_row[c] is not None:
                        val = next_row[c]
                        # Time data is either: string with ":" or float < 1 (Excel time serial)
                        if isinstance(val, str) and ':' in val:
                            has_time_data = True
                            break
                        elif isinstance(val, (int, float)) and 0 < val < 1:
                            has_time_data = True
                            break
                if has_time_data:
                    attendance_row = next_row
                    break
            
            if not attendance_row:
                continue  # No attendance data found for this employee
            
            # 🔑 Parse attendance for each day from the CORRECT row
            for day in range(1, days_in_month + 1):
                col_idx = next((c for c, d in date_columns.items() if d == day), None)
                if col_idx is None or col_idx >= len(attendance_row):
                    continue
                
                cell_value = attendance_row[col_idx]
                date_obj = date(current_year, current_month, day)
                
                # Parse times from cell
                times = parse_times_from_cell(cell_value)
                
                if times:
                    times.sort()
                    # Single time = Time In only; Multiple = First=In, Last=Out
                    time_in = times[0]
                    time_out = times[-1] if len(times) > 1 else None
                else:
                    # 🔑 No attendance = explicitly None (not midnight!)
                    time_in = None
                    time_out = None
                
                attendance_records.append({
                    'user_id': current_employee['user_id'],
                    'employee_name': current_employee['employee_name'],
                    'department': current_employee['department'],
                    'date': date_obj.isoformat(),
                    'date_display': date_obj.strftime('%Y-%m-%d'),
                    'time_in': time_in.isoformat() if time_in else None,
                    'time_out': time_out.isoformat() if time_out else None,
                })
    
    # 🔑 Sort by user_id first, then date ascending (1→29)
    attendance_records.sort(key=lambda x: (x['user_id'], x['date']))
    
    return attendance_records


def parse_times_from_cell(cell_value):
    """Extract all time values from a cell - returns list of datetime.time objects"""
    import re
    from datetime import datetime, time
    
    times = []
    
    # Handle None/empty
    if not cell_value:
        return times
    
    cell_str = str(cell_value).strip()
    if not cell_str or cell_str.lower() in ['none', 'null', '', '0', '0.0']:
        return times
    
    # 🔑 Handle Excel time serial numbers (floats between 0 and 1)
    if isinstance(cell_value, (int, float)):
        # Excel: 1.0 = 24 hours, 0.5 = 12:00, 0.9 = 21:36
        # Only process if it's actually a time (fraction < 1)
        if 0 < cell_value < 1:
            try:
                total_seconds = int(cell_value * 86400)  # 86400 seconds/day
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                # 🔑 Don't return midnight (00:00:00) as valid attendance
                if hours == 0 and minutes == 0:
                    return []
                return [time(hours % 24, minutes, seconds)]
            except:
                return []
        else:
            return []  # It's a day number or invalid
    
    # 🔑 Handle string times (e.g., "21:36", "21:36\n21:43")
    # Split by newlines, commas, or multiple spaces
    time_strings = re.split(r'\n|\r|,\s*|\s{2,}', cell_str)
    
    for time_str in time_strings:
        time_str = time_str.strip()
        if time_str and time_str.lower() not in ['none', 'null', '']:
            parsed = parse_time(time_str)
            if parsed:
                times.append(parsed)
    
    return times  # List of time objects (or empty list)


def parse_time(time_str):
    """Parse time string to datetime.time object"""
    from datetime import datetime
    import re
    
    if not time_str:
        return None
    
    try:
        time_str = str(time_str).strip().upper()
        time_str = re.sub(r'\s+', ' ', time_str)  # Normalize spaces
        
        # Try common time formats
        formats = [
            '%H:%M:%S',     # 21:36:00
            '%H:%M',        # 21:36
            '%I:%M:%S %P',  # 9:36:00 pm
            '%I:%M %P',     # 9:36 pm
            '%I:%M%P',      # 9:36pm
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(time_str, fmt).time()
            except ValueError:
                continue
        
        # 🔑 Fallback: extract HH:MM with regex
        match = re.match(r'(\d{1,2}):(\d{2})', time_str)
        if match:
            h, m = int(match.group(1)), int(match.group(2))
            if 0 <= h <= 23 and 0 <= m <= 59:
                # Don't return midnight as valid attendance time
                if h == 0 and m == 0:
                    return None
                return time(h, m)
        
        return None
    except:
        return None


@hr_admin_bp.route('/attendance/<int:attendance_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_attendance(attendance_id):
    attendance = Attendance.query.get_or_404(attendance_id)
    
    if request.method == 'POST':
        # Get form data
        time_in_str = request.form.get('time_in')  # e.g., "08:54"
        time_out_str = request.form.get('time_out')  # e.g., "17:00"
        status = request.form.get('modal_status')
        remarks = request.form.get('remarks', '')

        # Convert strings to datetime.time objects if not empty
        if time_in_str:
            h, m = map(int, time_in_str.split(":"))
            attendance.time_in = time(hour=h, minute=m)
        else:
            attendance.time_in = None

        if time_out_str:
            h, m = map(int, time_out_str.split(":"))
            attendance.time_out = time(hour=h, minute=m)
        else:
            attendance.time_out = None

        # Update other fields
        attendance.status = status
        attendance.remarks = remarks

        # Recalculate working hours
        attendance.calculate_working_hours()

        # Commit changes
        db.session.commit()
        flash('Attendance updated!', 'success')
        return redirect(url_for('hr_admin_bp.view_attendance'))

    # GET request for modal JSON
    if request.headers.get('Accept') == 'application/json':
        return {
            "attendance_id": attendance.id,
            "date": attendance.date.strftime('%Y-%m-%d'),
            "time_in": attendance.time_in.strftime('%H:%M') if attendance.time_in else '',
            "time_out": attendance.time_out.strftime('%H:%M') if attendance.time_out else '',
            "modal_status": attendance.status,
            "remarks": attendance.remarks or ''
        }

    # fallback to page (optional)
    return render_template('hr/admin/atttendance/edit_attendance.html', attendance=attendance)



@hr_admin_bp.route('/add_manual_attendance', methods=['POST'])
@admin_required
@login_required  
def add_manual_attendance():
    try:
        # ===== Get form data =====
        employee_id = request.form.get('employee_id')
        date_str = request.form.get('date')
        time_in_str = request.form.get('time_in', '').strip()
        time_out_str = request.form.get('time_out', '').strip()
        status = request.form.get('status')
        remarks = request.form.get('remarks', '').strip()

        # ===== Validate required fields =====
        if not employee_id or not date_str or not status:
            flash("Employee, Date, and Status are required.", "error")
            return redirect(url_for('hr_admin_bp.view_attendance'))

        # ===== Validate employee exists and is active =====
        emp = Employee.query.filter_by(id=int(employee_id), archived=False).first()
        if not emp:
            flash("Employee not found or archived.", "error")
            return redirect(url_for('hr_admin_bp.view_attendance'))

        # ===== Parse date =====
        try:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            flash("Invalid date format. Use YYYY-MM-DD.", "error")
            return redirect(url_for('hr_admin_bp.view_attendance'))

        # ===== Parse times safely =====
        time_in_obj = None
        time_out_obj = None
        
        if time_in_str:
            try:
                h, m = map(int, time_in_str.split(":"))
                if 0 <= h <= 23 and 0 <= m <= 59:
                    time_in_obj = time(hour=h, minute=m)
            except Exception:
                flash(f"Invalid time_in format: {time_in_str}", "error")
                return redirect(url_for('hr_admin_bp.view_attendance'))

        if time_out_str:
            try:
                h, m = map(int, time_out_str.split(":"))
                if 0 <= h <= 23 and 0 <= m <= 59:
                    time_out_obj = time(hour=h, minute=m)
            except Exception:
                flash(f"Invalid time_out format: {time_out_str}", "error")
                return redirect(url_for('hr_admin_bp.view_attendance'))

        # ===== Check for duplicate =====
        existing_att = Attendance.query.filter_by(
            employee_id=emp.id, 
            date=attendance_date
        ).first()
        
        if existing_att:
            flash(f"Attendance already exists for {emp.get_full_name()} on {attendance_date}.", "error")
            return redirect(url_for('hr_admin_bp.view_attendance'))

        # ✅ CREATE ATTENDANCE WITH KEYWORD ARGUMENTS (FIXES THE ERROR)
        new_attendance = Attendance(
            employee_id=emp.id,           # ✅ keyword
            date=attendance_date,          # ✅ keyword
            time_in=time_in_obj,           # ✅ keyword
            time_out=time_out_obj,         # ✅ keyword
            status=status,                 # ✅ keyword
            remarks=remarks if remarks else None  # ✅ keyword
        )

        db.session.add(new_attendance)
        db.session.commit()  # ✅ Commit first

        # 🎯 Defer late computation until AFTER this request completes
        @after_this_request
        def compute_late_after_response(response):
            try:
                late_data = extract_late_from_attendance(new_attendance)
                if late_data:
                    existing = LateComputation.query.filter_by(attendance_id=new_attendance.id).first()
                    if existing:
                        existing.late_hours = late_data["late_hours"]
                        existing.late_minutes = late_data["late_minutes"]
                        existing.day_equivalent = late_data["day_equivalent"]
                        existing.remarks = "Updated from manual entry"
                    else:
                        record = LateComputation(
                            employee_id=new_attendance.employee_id,
                            attendance_id=new_attendance.id,
                            date=new_attendance.date,
                            late_days=0,
                            late_hours=late_data["late_hours"],
                            late_minutes=late_data["late_minutes"],
                            day_equivalent=late_data["day_equivalent"],
                            remarks="Auto-generated from manual attendance"
                        )
                        db.session.add(record)
                        db.session.commit()
            except Exception as e:
                current_app.logger.error(f"Late computation failed: {e}")
                db.session.rollback()
            return response  # Must return response

        flash(f"✅ Attendance for {emp.get_full_name()} on {attendance_date} added successfully.", "success")
        return redirect(url_for('hr_admin_bp.view_attendance'))

    except IntegrityError:
        db.session.rollback()
        flash("A record with this employee and date already exists.", "error")
        return redirect(url_for('hr_admin_bp.view_attendance'))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Add attendance error: {str(e)}", exc_info=True)
        flash("An unexpected error occurred. Please try again.", "error")
        return redirect(url_for('hr_admin_bp.view_attendance'))





@hr_admin_bp.route('/download/attendance-template')
@login_required
@admin_required
def download_attendance_template():
    # Define the template filename
    file_name = '1_(October)Employee Attendance Record.xls'
    
    # Build a portable path relative to your app's root directory
    file_path = os.path.join(current_app.root_path, 'templates', 'documents', file_name)

    # Fallback if the file is missing
    if not os.path.exists(file_path):
        flash('Attendance template file not found. Please contact support.', 'danger')
        # Replace 'hr_admin_bp.index' with your actual dashboard/home route
        return redirect(url_for('hr_admin_bp.index'))

    # Serve the file for download
    return send_file(
        file_path,
        as_attachment=True,
        download_name='Employee_Attendance_Record_Template.xls',
        mimetype='application/vnd.ms-excel'
    )